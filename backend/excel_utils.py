from datetime import timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from fastapi import HTTPException
from urls import get_holidays


# Convert weekday index to Korean name
def convert_day(day):
    days = ["월", "화", "수", "목", "금", "토", "일"]
    return days[day]


def generate_excel(start_date, end_date, repeat_num):
    national_holidays = get_holidays(start_date, end_date)
    difference = (end_date - start_date).days
    file = Workbook()

    # Create list of dates from start_date to end_date
    datelist = []
    for i in range(difference + 1):
        date = start_date + datetime.timedelta(days=i)
        day = convert_day(date.weekday())
        formatted_date = date.strftime("20%y년 %m월 %d일") + f"({day})"
        datelist.append(formatted_date)

    # Create worksheets based on repeat_num
    for n in range(1, repeat_num + 1):
        sheet = file.create_sheet(f"{n}개")
        sheet.column_dimensions["A"].width = 20

        all_dates = [date for date in datelist for _ in range(n)]

        # Populate sheet with dates
        for idx, item in enumerate(all_dates, start=1):
            sheet.append([item])  # Add date to Excel

            cell = sheet[f"A{idx}"]

            # Highlight Saturdays in red
            if "토" in item:
                cell.fill = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")

            # Highlight national holidays in orange
            if item.split("(")[0] in national_holidays:
                cell.fill = PatternFill(start_color="f5c77e", end_color="f5c77e", fill_type="solid")

    # Define the save directory
    save_dir = "temp_file"
    os.makedirs(save_dir, exist_ok=True)  # Create directory if it doesn't exist

    # Save the file
    filename = f"{start_date.month}월 - {end_date.month}월.xlsx"
    full_path = os.path.join(save_dir, filename)

    file.save(full_path)
    file.close()

    return full_path  # Return the path
